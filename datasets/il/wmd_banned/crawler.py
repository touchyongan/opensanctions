from zavod import Context
from typing import List, Optional, Tuple
from rigour.mime.types import XLSX
import openpyxl
from zavod import helpers as h
import shutil
import re
import datetime


from typing import Dict, Generator, Optional, Union
from datetime import datetime
from normality import slugify, stringify
from xlrd.book import Book  # type: ignore
from xlrd.sheet import Cell  # type: ignore
from xlrd.xldate import xldate_as_datetime  # type: ignore
from nomenklatura.util import datetime_iso
from openpyxl.worksheet.worksheet import Worksheet

from zavod.context import Context


def parse_xlsx_sheet(
    context: Context,
    sheet: Worksheet,
    skiprows: int = 0,
    header_lookup: Optional[str] = None,
) -> Generator[Dict[str, str | None], None, None]:
    """
    Parse an Excel sheet into a sequence of dictionaries.

    Args:
        context: Crawler context.
        sheet: The Excel sheet.
        skiprows: The number of rows to skip.
        header_lookup: The lookup key for translating headers.
    """
    headers = None
    row_counter = 0

    for row in sheet.iter_rows():
        # Increment row counter
        row_counter += 1

        # Skip the desired number of rows
        if row_counter <= skiprows:
            continue
        cells = [c.value for c in row]
        if headers is None:
            headers = []
            for idx, header in enumerate(cells):
                if header is None:
                    header = f"column_{idx}"
                if header_lookup:
                    header = context.lookup_value(
                        header_lookup,
                        stringify(header),
                        stringify(header),
                    )
                headers.append(slugify(header, sep="_"))
            continue

        record = {}
        for header, value in zip(headers, cells):
            if isinstance(value, datetime):
                value = value.date()
            record[header] = stringify(value)
        if len(record) == 0:
            continue
        if all(v is None for v in record.values()):
            continue
        yield record


def extract_passport_no(text: str):
    """
    Extract passport numbers from a given text.
    """
    if not text:
        return None
    text = str(text)
    pattern = r"\b[A-Z0-9]{5,}\b"
    matches = re.findall(pattern, text)

    return matches


def extract_n_pop_address(text: str):
    """
    Extract address and update the text by removing the extracted address.
    """
    if not text:
        return None, None

    pattern = r"(address|location):\s*(.*)"
    match = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
    if match:
        address = match.group(2).strip()

        # Remove the extracted location from the original text
        updated_text = re.sub(
            pattern, "", text, flags=re.DOTALL | re.IGNORECASE
        ).strip()

        return address, updated_text

    else:
        return None, text


def clean_date(date_str: str):
    """
    Clean the  date string by replacing newlines, colons, and dots with a
    space character and return a split of dates
    """
    if not date_str:
        return []
    date_str = str(date_str).lower()
    date_str = re.sub(r"[:\n\.]|dob|\xa0", " ", date_str).strip()

    # print("================")
    # print(date_str)
    # print(h.multi_split(date_str, ["a)", " b)", "c)", "d)", "e)"]))
    return h.multi_split(date_str, ["a)", " b)", "c)", "d)", "e)"])


def clean_name(name: str | List[str]):
    """
    Clean the given name (or list of names) by removing newlines, quotes, colons, and trailing dots
    """
    if isinstance(name, list):
        return [re.sub(r"[\n':]|^\.*|\.*$", "", match).strip() for match in name]

    return re.sub(r"[\n':]|^\.*|\.*$", "", name).strip()


def format_numbered_listings(text: str):
    """
    Extract items from a text that follows numbered or lettered listings.
    """
    if not text:
        return None
    patterns = [
        r"\b[a-z]\)\s(.*?)(?=\s[a-z]\)|$)",  # Matches 'a) text b) text'
        r"\b\d+:\s(.*?)(?=\s\d+:|$)",  # Matches '1: text 2: text'
    ]
    for pattern in patterns:
        matches = re.findall(pattern, text, re.DOTALL | re.VERBOSE | re.IGNORECASE)
        if matches:
            return matches

    else:
        return text


def sheet_to_tuple(sheet):
    """
    Convert rows from an Excel sheet to a tuple.
    Index of tuple is used to locate attributes in the crawler
    as the headers/keys are in hebrew
    """
    headers: Optional[List[str]] = None
    for row in sheet.iter_rows(min_row=0):
        cells = [c.value for c in row]

        if headers is None:
            if all(cells):
                headers = [h for h in cells]
            continue

        if len(cells) == len(headers):
            if 'מס"ד' in cells:  # cell that contains headers, skip it
                continue

            none_count = sum(1 for item in cells if not item)
            if (
                none_count > len(headers) // 2
            ):  #  cells that do not contain data up to half of the headers count, skip it as it is probably a subheader
                continue

            yield cells


def parse_sheet_row(context: Context, row: Tuple[str | int, ...]):
    assert (
        len(row) == 11
    ), "Crawler was developed with excel sheet of 11 columns, please check that attributes still conforms to index used"

    other_info = row.pop(10)
    address_nationality = row.pop(9)

    # Extract address if it exists from either the info or nationality attibrutes
    info_address, notes = extract_n_pop_address(other_info)
    nat_address, nationality = extract_n_pop_address(address_nationality)
    address = info_address or nat_address

    dob = row.pop(8)
    # dob = hebrew_to_datetime(dob)
    # dob = h.parse_date(dob, "%d %b. %Y")  #

    passport = row.pop(7)
    parse_name = row.pop(6)

    # p
    name_split = re.split(
        r"\bA\.K\.A\b|;|ידוע גם|AKA:", parse_name, flags=re.IGNORECASE
    )
    name = format_numbered_listings(name_split[0].strip())

    if isinstance(name, list) and len(name) > 1:
        alias = name[1:]
        name = name[0]
    else:
        alias = []

    other_alias = [format_numbered_listings(alias) for alias in name_split[1:]]
    flattened_alias = [
        item
        for sublist in other_alias
        for item in (sublist if isinstance(sublist, list) else [sublist])
    ]
    alias += flattened_alias

    name = clean_name(name)
    alias = clean_name(alias)

    isreal_adoption_date = row.pop(5)
    # isreal_adoption_date = hebrew_to_datetime(isreal_adoption_date)  # permanent

    isreal_temp_adoption_date = row.pop(4)
    # isreal_temp_adoption_date = hebrew_to_datetime(isreal_temp_adoption_date)

    serial_no = row.pop(3)
    originally_declared_by = row.pop(2)
    declaration_date = row.pop(1)
    # if declaration_date:
    #     declaration_date = h.parse_date(declaration_date, "%d %b. %Y")

    record_id = row.pop(0)

    if "iri" in serial_no.lower() or "pi" in serial_no.lower():
        entity = context.make("Person")
        entity.id = context.make_id("Person", f"{record_id}-{serial_no}")
        entity.add("passportNumber", extract_passport_no(passport))
        entity.add("nationality", nationality)

        h.apply_dates(
            entity,
            "birthDate",
            clean_date(dob),
        )

    elif "ire" in serial_no.lower() or "pe" in serial_no.lower():
        entity = context.make("Organization")
        entity.id = context.make_id("Company", f"{record_id}-{serial_no}")

    else:
        context.log.warn(f"Entity not recognized from serial number: {serial_no}")

    entity.add("name", name)
    entity.add("alias", alias)
    entity.add("notes", notes)
    entity.add("address", format_numbered_listings(address))

    sanction = h.make_sanction(context, entity)
    sanction.add("authority", originally_declared_by, lang="he")

    h.apply_dates(sanction, "listingDate", clean_date(declaration_date))

    h.apply_dates(
        sanction,
        "startDate",
        clean_date(isreal_adoption_date),
    )

    context.emit(entity, target=True)
    context.emit(sanction)


def crawl(context: Context):
    assert context.dataset.base_path is not None

    data_path = context.dataset.base_path / "data.xlsx"
    source_path = context.get_resource_path("source.xlsx")

    shutil.copyfile(data_path, source_path)

    context.export_resource(source_path, XLSX, title=context.SOURCE_TITLE)
    rows = sheet_to_tuple(
        openpyxl.load_workbook(source_path, read_only=True).worksheets[0]
    )

    # for row in rows:
    #     parse_sheet_row(context, row)

    #
    for row in parse_xlsx_sheet(
        context,
        openpyxl.load_workbook(source_path, read_only=True).worksheets[0],
        header_lookup="columns",
    ):
        print(row)
